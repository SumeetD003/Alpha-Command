import time, concurrent.futures, os, sys
# Add the parent directory of TradeMaster to sys.path
sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), '..', '..')))
import pandas as pd
from datetime import datetime
from TradeMaster.data_reader.data_reader import DataReader
from TradeMaster.backtesting import Backtest
from datetime import datetime
import pandas as pd
import numpy as np
import itertools, math
import matplotlib.pyplot as plt
from openpyxl import Workbook
from openpyxl.drawing.image import Image

class HyperParameterOptimizer:
    def __init__(self, strategy, * ,
                 cash: float = 10_000,
                 holding: dict = {},
                 commission: float = .0,
                 margin: float = 1.,
                 trade_on_close=False,
                 hedging=False,
                 exclusive_orders=False,
                 trade_start_date=None,
                 lot_size=1,
                 fail_fast=True,
                 storage: dict | None = None,
                 is_option: bool = False ):
        '''
        Take as input all the backtesting parameters at time of initialisation
        '''
        self.strategy = strategy
        self.cash = cash
        self.commission = commission
        self.holding = holding
        self.margin = margin
        self.trade_on_close = trade_on_close
        self.hedging = hedging
        self.exclusive_orders = exclusive_orders
        self.trade_start_date = trade_start_date
        self.lot_size= lot_size
        self.fail_fast = fail_fast
        self.storage = storage
        self.is_option = is_option 
        self.reader = DataReader()
        self.reader.initialize(host='qdb.satyvm.com', port=443, https = True, username='2Cents', password='2Cents$101')
        self.num_processes = 5      
          


    def generate_heatmaps(self,
                            top_sharpe_df,
                            top_equity_df,
                            top_winrate_df,
                            sharpe_df,
                            equity_df,
                            winrate_df,
                            output_file,
                            png_directory,
                            stock
                        ):
        """
        Create an Excel file with separate sheets for Sharpe Ratio, Equity Final, and Win Rate.
        Each sheet includes a top 10 percentile dataframe as a table, followed by heatmaps.

        Args:
            top_sharpe_df (pd.DataFrame): Top 10 percentile dataframe for Sharpe Ratio.
            top_equity_df (pd.DataFrame): Top 10 percentile dataframe for Equity Final.
            top_winrate_df (pd.DataFrame): Top 10 percentile dataframe for Win Rate.
            sharpe_df (pd.DataFrame): Complete dataframe for heatmap generation (Sharpe Ratio).
            equity_df (pd.DataFrame): Complete dataframe for heatmap generation (Equity Final).
            winrate_df (pd.DataFrame): Complete dataframe for heatmap generation (Win Rate).
            output_file (str): Path to save the Excel file.
        """

        # Metric-specific data
        metrics_data = [
            ("Sharpe Ratio", top_sharpe_df, sharpe_df),
            ("Equity Final [$]", top_equity_df, equity_df),
            ("Win Rate [%]", top_winrate_df, winrate_df),
        ]

        # Initialize Excel workbook
        workbook = Workbook()

        for metric_name, top_df, full_df in metrics_data:
            # Create a new sheet for the metric
            sheet_name = metric_name.split("[", 1)[0]
            sheet = workbook.create_sheet(title=sheet_name)

            # Write the top 10 percentile dataframe as a table
            for row_idx, row in enumerate(
                [top_df.columns.tolist()] + top_df.values.tolist(), start=1
            ):
                for col_idx, value in enumerate(row, start=1):
                    sheet.cell(row=row_idx, column=col_idx, value=value)

            # Add some space before the heatmaps
            start_row = len(top_df) + 3

            # Extract parameter columns from the dataframe
            parameter_columns = [
                col for col in full_df.columns if col not in [metric_name]
            ]

            # Create all possible parameter pairs
            parameter_pairs = list(itertools.combinations(parameter_columns, 2))

            # Iterate over parameter pairs and create heatmaps
            for param_x, param_y in parameter_pairs:
                # Group data by the parameter pair and calculate average metric
                grouped = (
                    full_df.groupby([param_x, param_y])[metric_name]
                    .mean()
                    .unstack(fill_value=np.nan)  # Pivot table to create a matrix
                )

                # Set up colormap to display NaN values in white
                cmap = plt.cm.viridis
                cmap = cmap.copy()
                cmap.set_bad(color="white")

                # Create the heatmap
                plt.figure(figsize=(5, 4))  # Resize to half the original size
                ax = plt.gca()
                heatmap = ax.imshow(
                    grouped.values, cmap=cmap, aspect="auto", interpolation="nearest"
                )

                # Set axis labels and ticks
                ax.set_xticks(np.arange(len(grouped.columns)))
                ax.set_yticks(np.arange(len(grouped.index)))
                ax.set_xticklabels(grouped.columns, rotation=45)
                ax.set_yticklabels(grouped.index)
                ax.set_xlabel(param_y)
                ax.set_ylabel(param_x)
                ax.set_title(f"Heatmap of {metric_name}: {param_x} vs {param_y}")

                # Add a colorbar
                plt.colorbar(heatmap, ax=ax)

                # Save heatmap as an image
                image_file = f"{param_x}_{param_y}_{metric_name}_{stock}_heatmap.png"
                image_file = os.path.join(png_directory, image_file)
                plt.savefig(image_file, bbox_inches="tight", dpi=150)
                plt.close()

                # Embed the image in the Excel sheet
                img = Image(image_file)
                img.anchor = f"A{start_row}"  # Anchor at the appropriate row
                sheet.add_image(img)

                # Move the start row down for the next heatmap
                start_row += 30  # Adjust for vertical spacing
                
        if "Sheet" in workbook.sheetnames:
            workbook.remove(workbook["Sheet"])
        # Save the Excel workbook
        workbook.save(output_file)
        print(f"Metrics and heatmaps embedded in {output_file}")

    def delete_png_files(self, png_directory):
        for filename in os.listdir(png_directory):
            if filename.endswith(".png"):
                file_path = os.path.join(png_directory, filename)
                os.remove(file_path) 
    
    def optimize_stock(self, 
                    stock, 
                    timeframe, 
                    market, 
                    exchange = None,
                    * ,
                    method = 'grid',
                    max_tries = None,
                    constraint = None,
                    random_state = None,
                    **kwargs
                    ):
        try:
            data = self.reader.fetch_stock(stock, timeframe, market, exchange) # Get the dataframe for stock
        except Exception as e:
            print(f"Error while fetching data for {stock}: {e}")
            raise
        
        print("Data fetched for",stock)
        
        bt = Backtest(data, self.strategy, 
                    cash = self.cash,
                    commission = self.commission,
                    holding = self.holding,
                    margin = self.margin,
                    trade_on_close = self.trade_on_close,
                    hedging = self.hedging,
                    exclusive_orders = self.exclusive_orders,
                    trade_start_date = self.trade_start_date,
                    lot_size= self.lot_size,
                    fail_fast = self.fail_fast,
                    storage = self.storage,
                    is_option = self.is_option 
                    )
        sharpe_ratio_best_results, sharpe_ratio_heatmap = bt.optimize(maximize = 'Sharpe Ratio',
                                                                      method = method,
                                                                      max_tries = max_tries,
                                                                      constraint = constraint,
                                                                      return_heatmap = True,
                                                                      random_state = random_state,
                                                                      **kwargs
                                                                      )
        
        win_rate_best_results, win_rate_heatmap = bt.optimize(maximize = 'Win Rate [%]',
                                                                      method = method,
                                                                      max_tries = max_tries,
                                                                      constraint = constraint,
                                                                      return_heatmap = True,
                                                                      random_state = random_state,
                                                                      **kwargs
                                                                      )
        
        equity_final_best_results, equity_final_heatmap = bt.optimize(maximize = 'Equity Final [$]',
                                                                      method = method,
                                                                      max_tries = max_tries,
                                                                      constraint = constraint,
                                                                      return_heatmap = True,
                                                                      random_state = random_state,
                                                                      **kwargs
                                                                      )
        
        
        # Create output directory
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        output_directory_path = os.path.join(self.get_root_dir(), f'hyperparameter_optimizer_output_{timestamp}')
        os.makedirs(output_directory_path)
        
        sharpe_ratio_df = sharpe_ratio_heatmap.reset_index()
        win_rate_df = win_rate_heatmap.reset_index()
        equity_final_df = equity_final_heatmap.reset_index()
        
        # sharpe_ratio_top_10p_df = sharpe_ratio_df[sharpe_ratio_df["Sharpe Ratio"] >= sharpe_ratio_df["Sharpe Ratio"].quantile(0.90)].sort_values(by="Sharpe Ratio", ascending=False)
        # win_rate_ratio_top_10p_df = win_rate_df[win_rate_df["Win Rate [%]"] >= win_rate_df["Win Rate [%]"].quantile(0.90)].sort_values(by="Win Rate [%]", ascending=False)
        # equity_final_top_10p_df = equity_final_df[equity_final_df["Equity Final [$]"] >= equity_final_df["Equity Final [$]"].quantile(0.90)].sort_values(by="Equity Final [$]", ascending=False)
        
        sharpe_ratio_sorted_df = sharpe_ratio_df.sort_values(by="Sharpe Ratio", ascending=False)
        equity_final_sorted_df = equity_final_df.sort_values(by="Equity Final [$]", ascending=False)
        win_rate_sorted_df = win_rate_df.sort_values(by="Win Rate [%]", ascending=False)
        
        sheet_path = os.path.join(output_directory_path, f'summary_{stock}.xlsx')

        self.generate_heatmaps(sharpe_ratio_sorted_df,equity_final_sorted_df,win_rate_sorted_df,
                               sharpe_ratio_df, equity_final_df, win_rate_df, sheet_path, output_directory_path, stock)
        self.delete_png_files(output_directory_path)
        
    def optimize_universe_worker(self, data_chunk):
        for data_tuple in data_chunk:
            stock, data, method, max_tries, constraint, random_state, kwargs = data_tuple
            bt = Backtest(data, self.strategy, 
                        cash = self.cash,
                        commission = self.commission,
                        holding = self.holding,
                        margin = self.margin,
                        trade_on_close = self.trade_on_close,
                        hedging = self.hedging,
                        exclusive_orders = self.exclusive_orders,
                        trade_start_date = self.trade_start_date,
                        lot_size= self.lot_size,
                        fail_fast = self.fail_fast,
                        storage = self.storage,
                        is_option = self.is_option 
                        )
            sharpe_ratio_best_results, sharpe_ratio_heatmap = bt.optimize(maximize = 'Sharpe Ratio',
                                                                        method = method,
                                                                        max_tries = max_tries,
                                                                        constraint = constraint,
                                                                        return_heatmap = True,
                                                                        random_state = random_state,
                                                                        **kwargs
                                                                        )
            
            win_rate_best_results, win_rate_heatmap = bt.optimize(maximize = 'Win Rate [%]',
                                                                        method = method,
                                                                        max_tries = max_tries,
                                                                        constraint = constraint,
                                                                        return_heatmap = True,
                                                                        random_state = random_state,
                                                                        **kwargs
                                                                        )
            
            equity_final_best_results, equity_final_heatmap = bt.optimize(maximize = 'Equity Final [$]',
                                                                        method = method,
                                                                        max_tries = max_tries,
                                                                        constraint = constraint,
                                                                        return_heatmap = True,
                                                                        random_state = random_state,
                                                                        **kwargs
                                                                        )
            
            sharpe_ratio_df = sharpe_ratio_heatmap.reset_index()
            win_rate_df = win_rate_heatmap.reset_index()
            equity_final_df = equity_final_heatmap.reset_index()
            
            # sharpe_ratio_top_10p_df = sharpe_ratio_df[sharpe_ratio_df["Sharpe Ratio"] >= sharpe_ratio_df["Sharpe Ratio"].quantile(0.90)].sort_values(by="Sharpe Ratio", ascending=False)
            # win_rate_ratio_top_10p_df = win_rate_df[win_rate_df["Win Rate [%]"] >= win_rate_df["Win Rate [%]"].quantile(0.90)].sort_values(by="Win Rate [%]", ascending=False)
            # equity_final_top_10p_df = equity_final_df[equity_final_df["Equity Final [$]"] >= equity_final_df["Equity Final [$]"].quantile(0.90)].sort_values(by="Equity Final [$]", ascending=False)
            
            sharpe_ratio_sorted_df = sharpe_ratio_df.sort_values(by="Sharpe Ratio", ascending=False)
            equity_final_sorted_df = equity_final_df.sort_values(by="Equity Final [$]", ascending=False)
            win_rate_sorted_df = win_rate_df.sort_values(by="Win Rate [%]", ascending=False)

            sheet_path = os.path.join(self.output_directory_path, f'summary_{stock}.xlsx')
            self.generate_heatmaps(sharpe_ratio_sorted_df,equity_final_sorted_df,win_rate_sorted_df,
                                sharpe_ratio_df, equity_final_df, win_rate_df, sheet_path, self.output_directory_path, stock)
    
    def optimize_universe(self, 
                        universe, 
                        timeframe, 
                        exchange = None,
                        * ,
                        method = 'grid',
                        max_tries = None,
                        constraint = None,
                        random_state = None,
                        **kwargs
                        ):
        data_all = self.reader.fetch_universe(universe, timeframe, exchange)
        print("Data fetched")
        data_tuples = [(stock, data, method,
                        max_tries, constraint, random_state, kwargs) for (stock, data) in data_all.items()]
        data_chunks = list(self.split_list(data_tuples, self.num_processes))
        
        # Create output directory
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        output_directory_path = os.path.join(self.get_root_dir(), f'hyperparameter_optimizer_output_{timestamp}')
        os.makedirs(output_directory_path)
        self.output_directory_path = output_directory_path
        
        start_time = time.time()

        with concurrent.futures.ProcessPoolExecutor(max_workers=self.num_processes) as executor:
            # Map the tasks to the stock name for later on
            futures = [executor.submit(self.optimize_universe_worker, data_chunk) for data_chunk in data_chunks]
            
            for future in concurrent.futures.as_completed(futures):
                future.result() 
                    
        self.delete_png_files(output_directory_path)            
                    
        end_time = time.time()
        print("Elapsed",end_time-start_time)
        
    
    def get_root_dir(self):
        return os.getcwd() 
    
    def split_list(self, data, n):
        chunk_size = math.ceil(len(data) / n)
        for i in range(0, len(data), chunk_size):
            yield data[i:i + chunk_size]
    
if __name__ == '__main__':
    print("Here in main")
    # bt = HyperParameterOptimizer()